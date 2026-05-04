"""
generate_report.py — Arb Bot Trade Report Generator

Reads from data/arb_positions.db + data log files and produces a
fully-formatted Excel workbook: arb_trade_report.xlsx

Sheets:
  1. Trade Log      — every trade, all fields, colour-coded P&L
  2. Open Positions — live open trades
  3. P&L Summary    — KPI dashboard with risk metrics
  4. Daily P&L      — day-by-day breakdown + cumulative
  5. Sessions       — bot run history parsed from DB + logs
  6. Scan Activity  — log-file scan stats (markets, opps, candidates)
  7. How to Refresh — usage instructions

Usage (run from arb-bot/):
    python scripts/generate_report.py [--output PATH] [--demo]

--demo   Injects sample trades so you can see how each sheet looks
         before the first real trade fires. Demo rows are clearly labelled.
"""

import argparse
import os
import re
import sqlite3
import sys
from datetime import datetime, timezone, timedelta
from typing import Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel

# ─────────────────────────────────────────────────────────────────────────────
# Colour palette
# ─────────────────────────────────────────────────────────────────────────────
C_DARK_BG    = "1A1A2E"   # dark navy header background
C_MID_BG     = "16213E"   # slightly lighter accent
C_ACCENT     = "0F3460"   # blue accent
C_GOLD       = "E94560"   # red-gold highlight
C_WHITE      = "FFFFFF"
C_LIGHT_GREY = "F5F5F5"
C_BORDER     = "D0D0D0"

C_GREEN_BG   = "E8F5E9"
C_GREEN_FG   = "1B5E20"
C_RED_BG     = "FFEBEE"
C_RED_FG     = "B71C1C"
C_AMBER_BG   = "FFF8E1"
C_AMBER_FG   = "E65100"
C_BLUE_FG    = "0000FF"   # hardcoded inputs (industry standard)
C_BLACK_FG   = "000000"   # formulas

FONT_NAME = "Arial"


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def hdr_font(bold=True, color=C_WHITE, size=10):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)

def body_font(bold=False, color=C_BLACK_FG, size=10):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)

def blue_font(size=10):
    """Industry standard: blue = hardcoded input."""
    return Font(name=FONT_NAME, color=C_BLUE_FG, size=size)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def right_align():
    return Alignment(horizontal="right", vertical="center")

def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def style_header_row(ws, row: int, bg: str = C_DARK_BG, fg: str = C_WHITE,
                     height: float = 22):
    ws.row_dimensions[row].height = height
    for cell in ws[row]:
        if cell.value is not None or True:
            cell.font      = hdr_font(color=fg)
            cell.fill      = fill(bg)
            cell.alignment = center()
            cell.border    = thin_border()

def style_data_row(ws, row: int, alternate: bool = False):
    bg = C_LIGHT_GREY if alternate else C_WHITE
    ws.row_dimensions[row].height = 16
    for cell in ws[row]:
        cell.fill      = fill(bg)
        cell.font      = body_font()
        cell.alignment = left()
        cell.border    = thin_border()

def fmt_dt(iso_str: Optional[str]) -> str:
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return iso_str[:16] if iso_str else ""

def fmt_usd(val) -> str:
    if val is None:
        return ""
    return f"${val:,.2f}"

def pct(val) -> str:
    if val is None:
        return ""
    return f"{val:.2%}"

def calc_hold_hours(opened_at, closed_at) -> Optional[float]:
    if not opened_at or not closed_at:
        return None
    try:
        o = datetime.fromisoformat(opened_at.replace("Z", "+00:00"))
        c = datetime.fromisoformat(closed_at.replace("Z", "+00:00"))
        return round((c - o).total_seconds() / 3600, 1)
    except Exception:
        return None

def calc_est_fee(kalshi_price, kalshi_contracts, kalshi_fee_rate=0.07):
    """Estimated Kalshi taker fee = 7% × entry price × contracts."""
    if kalshi_price is None or kalshi_contracts is None:
        return None
    return round(kalshi_fee_rate * kalshi_price * kalshi_contracts, 4)

def annualised_return(edge_per_day) -> Optional[float]:
    if edge_per_day is None:
        return None
    return round(edge_per_day * 365, 4)

def max_drawdown(pnl_series):
    if not pnl_series:
        return 0.0
    cumulative, peak, max_dd = 0.0, 0.0, 0.0
    for p in pnl_series:
        cumulative += p
        peak = max(peak, cumulative)
        max_dd = max(max_dd, peak - cumulative)
    return round(max_dd, 4)

def sharpe(pnl_series):
    import math
    if len(pnl_series) < 2:
        return 0.0
    mean = sum(pnl_series) / len(pnl_series)
    variance = sum((p - mean) ** 2 for p in pnl_series) / (len(pnl_series) - 1)
    std = math.sqrt(variance)
    return round((mean / std * math.sqrt(365)) if std > 0 else 0.0, 3)


# ─────────────────────────────────────────────────────────────────────────────
# DB reader
# ─────────────────────────────────────────────────────────────────────────────

def load_data(db_path: str):
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    positions = [dict(r) for r in conn.execute(
        "SELECT * FROM positions ORDER BY opened_at DESC LIMIT 5000"
    ).fetchall()]

    fills = [dict(r) for r in conn.execute(
        "SELECT * FROM fills ORDER BY filled_at DESC LIMIT 10000"
    ).fetchall()]

    sessions = [dict(r) for r in conn.execute(
        "SELECT * FROM sessions ORDER BY started_at DESC LIMIT 200"
    ).fetchall()]

    conn.close()
    return positions, fills, sessions


# ─────────────────────────────────────────────────────────────────────────────
# Log parser
# ─────────────────────────────────────────────────────────────────────────────

def parse_scan_log(log_path: str) -> list:
    """
    Parse bot log file for per-scan stats.
    Returns list of dicts: {timestamp, scan_num, kalshi_mkts, poly_mkts,
                             candidates, opportunities, trades}
    """
    if not os.path.exists(log_path):
        return []

    scan_re    = re.compile(r"Scan #(\d+) \((\d{2}:\d{2}:\d{2} UTC)\)")
    markets_re = re.compile(r"Markets: (\d+) Kalshi.*?/ (\d+)")
    cands_re   = re.compile(r"Candidates?: (\d+)|Matched (\d+) candidate")
    opps_re    = re.compile(r"Opportunities?: (\d+)")
    trade_re   = re.compile(r"Trade #(\d+):.*cost=\$?([\d.]+)")
    date_re    = re.compile(r"^(\d{4}-\d{2}-\d{2})")  # for dated logs

    scans = []
    current = {}
    log_date = None

    try:
        with open(log_path, "r", errors="replace") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue

                # Try to grab date prefix (some logs have it)
                dm = date_re.match(line)
                if dm:
                    log_date = dm.group(1)

                m = scan_re.search(line)
                if m:
                    if current:
                        scans.append(current)
                    ts_str = m.group(2)
                    full_ts = f"{log_date or 'unknown'} {ts_str}" if log_date else ts_str
                    current = {
                        "scan_num": int(m.group(1)),
                        "timestamp": full_ts,
                        "kalshi_mkts": 0,
                        "poly_mkts": 0,
                        "candidates": 0,
                        "opportunities": 0,
                        "trades": 0,
                    }
                    continue

                if not current:
                    continue

                mm = markets_re.search(line)
                if mm:
                    current["kalshi_mkts"] = int(mm.group(1))
                    current["poly_mkts"]   = int(mm.group(2))

                cm = cands_re.search(line)
                if cm:
                    current["candidates"] = int(cm.group(1) or cm.group(2) or 0)

                om = opps_re.search(line)
                if om:
                    current["opportunities"] = max(
                        current["opportunities"], int(om.group(1))
                    )

                tm = trade_re.search(line)
                if tm:
                    current["trades"] += 1

        if current:
            scans.append(current)
    except Exception as e:
        print(f"  [warn] Log parse error: {e}")

    return scans


# ─────────────────────────────────────────────────────────────────────────────
# Demo data
# ─────────────────────────────────────────────────────────────────────────────

DEMO_POSITIONS = [
    {
        "id": 1, "mode": "paper",
        "opened_at": "2026-04-20T14:23:00+00:00",
        "closed_at": "2026-04-23T09:00:00+00:00",
        "status": "closed",
        "kalshi_ticker": "KXBTC-25APR23-B74000",
        "kalshi_title": "BTC above $74,000 by Apr 23?",
        "kalshi_side": "yes",
        "kalshi_price": 0.42, "kalshi_contracts": 5,
        "kalshi_fill_price": 0.425,
        "poly_token_id": "0xabc123", "poly_question": "Bitcoin above $74k on April 23?",
        "poly_side": "BUY", "poly_price": 0.51, "poly_size_usd": 2.55, "poly_fill_price": 0.515,
        "gross_cost": 4.675, "expected_profit": 0.35, "actual_profit": 0.32,
        "actual_profit_pct": 0.0685, "match_score": 0.92, "llm_verified": 1,
        "edge_per_day": 0.023, "days_to_resolution": 3, "close_reason": "resolved_yes",
        "notes": "[DEMO]",
    },
    {
        "id": 2, "mode": "paper",
        "opened_at": "2026-04-21T10:05:00+00:00",
        "closed_at": "2026-04-24T18:00:00+00:00",
        "status": "closed",
        "kalshi_ticker": "KXFED-MAY26-525",
        "kalshi_title": "Fed rate unchanged at May 2026 meeting?",
        "kalshi_side": "no",
        "kalshi_price": 0.35, "kalshi_contracts": 8,
        "kalshi_fill_price": 0.358,
        "poly_token_id": "0xdef456", "poly_question": "Fed holds rates in May 2026?",
        "poly_side": "BUY", "poly_price": 0.59, "poly_size_usd": 4.72, "poly_fill_price": 0.595,
        "gross_cost": 7.584, "expected_profit": 0.60, "actual_profit": 0.58,
        "actual_profit_pct": 0.0765, "match_score": 0.88, "llm_verified": 1,
        "edge_per_day": 0.026, "days_to_resolution": 3, "close_reason": "resolved_no",
        "notes": "[DEMO]",
    },
    {
        "id": 3, "mode": "paper",
        "opened_at": "2026-04-22T08:30:00+00:00",
        "closed_at": "2026-04-25T12:00:00+00:00",
        "status": "closed",
        "kalshi_ticker": "KXBTC-25APR25-B76000",
        "kalshi_title": "BTC above $76,000 by Apr 25?",
        "kalshi_side": "yes",
        "kalshi_price": 0.38, "kalshi_contracts": 6,
        "kalshi_fill_price": 0.385,
        "poly_token_id": "0xghi789", "poly_question": "Bitcoin above $76k Apr 25?",
        "poly_side": "BUY", "poly_price": 0.56, "poly_size_usd": 3.36, "poly_fill_price": 0.562,
        "gross_cost": 6.672, "expected_profit": 0.52, "actual_profit": -0.18,
        "actual_profit_pct": -0.027, "match_score": 0.85, "llm_verified": 1,
        "edge_per_day": 0.009, "days_to_resolution": 3, "close_reason": "resolved_no",
        "notes": "[DEMO]",
    },
    {
        "id": 4, "mode": "paper",
        "opened_at": "2026-04-25T16:00:00+00:00",
        "closed_at": None,
        "status": "open",
        "kalshi_ticker": "KXCPI-MAY26-3",
        "kalshi_title": "US CPI above 3% in May 2026?",
        "kalshi_side": "no",
        "kalshi_price": 0.44, "kalshi_contracts": 4,
        "kalshi_fill_price": 0.442,
        "poly_token_id": "0xjkl012", "poly_question": "CPI inflation stays above 3% May 2026?",
        "poly_side": "BUY", "poly_price": 0.52, "poly_size_usd": 2.08, "poly_fill_price": 0.522,
        "gross_cost": 3.848, "expected_profit": 0.22, "actual_profit": None,
        "actual_profit_pct": None, "match_score": 0.91, "llm_verified": 1,
        "edge_per_day": 0.020, "days_to_resolution": 7, "close_reason": None,
        "notes": "[DEMO]",
    },
    {
        "id": 5, "mode": "paper",
        "opened_at": "2026-04-26T09:15:00+00:00",
        "closed_at": None,
        "status": "open",
        "kalshi_ticker": "KXBTC-26APR30-B78000",
        "kalshi_title": "BTC above $78,000 by Apr 30?",
        "kalshi_side": "yes",
        "kalshi_price": 0.48, "kalshi_contracts": 3,
        "kalshi_fill_price": 0.483,
        "poly_token_id": "0xmno345", "poly_question": "Bitcoin above $78k April 30?",
        "poly_side": "BUY", "poly_price": 0.50, "poly_size_usd": 1.50, "poly_fill_price": 0.502,
        "gross_cost": 2.949, "expected_profit": 0.15, "actual_profit": None,
        "actual_profit_pct": None, "match_score": 0.87, "llm_verified": 1,
        "edge_per_day": 0.017, "days_to_resolution": 4, "close_reason": None,
        "notes": "[DEMO]",
    },
]

DEMO_SESSIONS = [
    {"id": 1, "mode": "paper", "started_at": "2026-04-20T10:00:00+00:00",
     "ended_at": "2026-04-20T22:00:00+00:00", "trades_executed": 2,
     "gross_pnl": 0.32, "fees_paid": 0.09, "net_pnl": 0.23},
    {"id": 2, "mode": "paper", "started_at": "2026-04-21T09:00:00+00:00",
     "ended_at": "2026-04-21T21:00:00+00:00", "trades_executed": 1,
     "gross_pnl": 0.58, "fees_paid": 0.14, "net_pnl": 0.44},
    {"id": 3, "mode": "paper", "started_at": "2026-04-22T08:30:00+00:00",
     "ended_at": "2026-04-22T20:30:00+00:00", "trades_executed": 1,
     "gross_pnl": -0.18, "fees_paid": 0.11, "net_pnl": -0.29},
    {"id": 4, "mode": "paper", "started_at": "2026-04-25T15:00:00+00:00",
     "ended_at": None, "trades_executed": 0, "gross_pnl": 0.0, "fees_paid": 0.0, "net_pnl": 0.0},
]


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1 — Trade Log
# ─────────────────────────────────────────────────────────────────────────────

TRADE_HEADERS = [
    "Trade ID",          # A
    "Mode",              # B
    "Status",            # C
    "Opened At (UTC)",   # D
    "Closed At (UTC)",   # E
    "Hold Time (hrs)",   # F
    "Market Title",      # G
    "Kalshi Ticker",     # H
    "K Side",            # I  Kalshi side YES/NO
    "K Entry Price",     # J
    "K Fill Price",      # K
    "K Contracts",       # L
    "P Question",        # M
    "P Side",            # N
    "P Entry Price",     # O
    "P Fill Price",      # P
    "P Size (USDC)",     # Q
    "Gross Cost ($)",    # R
    "Est. K Fee ($)",    # S  Kalshi 7% fee on price×contracts
    "Exp. Profit ($)",   # T
    "Exp. Edge %",       # U  expected_profit / gross_cost
    "Act. Profit ($)",   # V
    "Act. P&L %",        # W
    "Net P&L After Fee", # X  actual_profit - est_fee
    "Annualised Ret.",   # Y  edge_per_day × 365
    "Days to Res.",      # Z
    "Edge / Day",        # AA
    "Match Score",       # AB
    "LLM Verified",      # AC
    "Close Reason",      # AD
    "W/L",               # AE
    "Notes",             # AF
]

def build_trade_log(ws, positions: list, sheet_title: str = "Trade Log"):
    ws.title = sheet_title
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # Title banner
    ws.merge_cells("A1:AF1")
    title_cell = ws["A1"]
    title_cell.value = f"  {sheet_title}  ·  Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.font  = Font(name=FONT_NAME, bold=True, color=C_WHITE, size=12)
    title_cell.fill  = fill(C_DARK_BG)
    title_cell.alignment = left()
    ws.row_dimensions[1].height = 26

    # Column headers
    for col_idx, hdr in enumerate(TRADE_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=hdr)
    style_header_row(ws, 2, bg=C_ACCENT)

    col_widths = {
        "A": 9, "B": 7, "C": 8, "D": 16, "E": 16, "F": 11,
        "G": 34, "H": 26, "I": 7, "J": 12, "K": 12, "L": 11,
        "M": 34, "N": 7, "O": 12, "P": 12, "Q": 13,
        "R": 13, "S": 12, "T": 13, "U": 11,
        "V": 13, "W": 11, "X": 15, "Y": 13,
        "Z": 11, "AA": 11, "AB": 11, "AC": 12,
        "AD": 16, "AE": 6, "AF": 20,
    }
    set_col_widths(ws, col_widths)

    # Data rows
    demo_flag = not any(not (p.get("notes") or "").startswith("[DEMO]") for p in positions if positions)

    for row_idx, p in enumerate(positions, start=3):
        alt = (row_idx % 2 == 0)
        style_data_row(ws, row_idx, alternate=alt)

        est_fee   = calc_est_fee(p.get("kalshi_price"), p.get("kalshi_contracts"))
        hold_hrs  = calc_hold_hours(p.get("opened_at"), p.get("closed_at"))
        ann_ret   = annualised_return(p.get("edge_per_day"))
        wl        = ""
        if p.get("actual_profit") is not None:
            wl = "W" if p["actual_profit"] > 0 else ("L" if p["actual_profit"] < 0 else "—")

        exp_edge_pct = None
        if p.get("expected_profit") and p.get("gross_cost"):
            try:
                exp_edge_pct = p["expected_profit"] / p["gross_cost"]
            except ZeroDivisionError:
                pass

        net_pnl_after_fee = None
        if p.get("actual_profit") is not None and est_fee is not None:
            net_pnl_after_fee = round(p["actual_profit"] - est_fee, 4)

        row_data = [
            p.get("id"),
            p.get("mode", "paper").upper(),
            (p.get("status") or "").upper(),
            fmt_dt(p.get("opened_at")),
            fmt_dt(p.get("closed_at")),
            hold_hrs,
            p.get("kalshi_title") or p.get("kalshi_ticker"),
            p.get("kalshi_ticker"),
            (p.get("kalshi_side") or "").upper(),
            p.get("kalshi_price"),
            p.get("kalshi_fill_price"),
            p.get("kalshi_contracts"),
            p.get("poly_question") or p.get("poly_token_id"),
            (p.get("poly_side") or "").upper(),
            p.get("poly_price"),
            p.get("poly_fill_price"),
            p.get("poly_size_usd"),
            p.get("gross_cost"),
            est_fee,
            p.get("expected_profit"),
            exp_edge_pct,
            p.get("actual_profit"),
            p.get("actual_profit_pct"),
            net_pnl_after_fee,
            ann_ret,
            p.get("days_to_resolution"),
            p.get("edge_per_day"),
            p.get("match_score"),
            "YES" if p.get("llm_verified") else "NO",
            (p.get("close_reason") or "").replace("_", " ").title(),
            wl,
            p.get("notes") or "",
        ]

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            col_letter = get_column_letter(col_idx)

            # Number formatting
            if col_letter in ("J", "K", "O", "P"):   # prices
                cell.number_format = "0.000"
            elif col_letter in ("R", "S", "T", "V", "X"):  # dollar amounts
                cell.number_format = '$#,##0.00;($#,##0.00);"-"'
            elif col_letter in ("U", "W", "Y"):  # percentages
                cell.number_format = "0.00%;(0.00%);-"
            elif col_letter in ("AB",):  # match score
                cell.number_format = "0.00"
            elif col_letter == "Q":
                cell.number_format = '$#,##0.00'
            elif col_letter == "AA":
                cell.number_format = "0.000%"

        # Colour-code W/L column
        wl_cell  = ws.cell(row=row_idx, column=31)  # AE
        pnl_cell = ws.cell(row=row_idx, column=22)  # V — actual_profit
        if wl == "W":
            wl_cell.fill  = fill(C_GREEN_BG)
            wl_cell.font  = Font(name=FONT_NAME, bold=True, color=C_GREEN_FG, size=10)
            pnl_cell.font = Font(name=FONT_NAME, color=C_GREEN_FG, size=10)
        elif wl == "L":
            wl_cell.fill  = fill(C_RED_BG)
            wl_cell.font  = Font(name=FONT_NAME, bold=True, color=C_RED_FG, size=10)
            pnl_cell.font = Font(name=FONT_NAME, color=C_RED_FG, size=10)

        # Demo rows: subtle amber tint on Notes cell
        notes_cell = ws.cell(row=row_idx, column=32)
        if (p.get("notes") or "").startswith("[DEMO]"):
            notes_cell.fill = fill(C_AMBER_BG)
            notes_cell.font = Font(name=FONT_NAME, italic=True,
                                   color=C_AMBER_FG, size=9)

    # Totals row
    if positions:
        total_row = len(positions) + 3
        ws.row_dimensions[total_row].height = 18
        ws.cell(total_row, 1, value="TOTALS").font = hdr_font()
        ws.cell(total_row, 1).fill = fill(C_MID_BG)

        for col_idx in (18, 19, 20, 22, 24):  # R S T V X
            col_l = get_column_letter(col_idx)
            first_data = 3
            last_data  = total_row - 1
            cell = ws.cell(total_row, col_idx,
                           value=f"=SUM({col_l}{first_data}:{col_l}{last_data})")
            cell.font   = hdr_font()
            cell.fill   = fill(C_MID_BG)
            nf = '$#,##0.00;($#,##0.00);"-"'
            cell.number_format = nf

    # Demo notice
    if any((p.get("notes") or "").startswith("[DEMO]") for p in positions):
        notice_row = (len(positions) + 5) if positions else 5
        ws.merge_cells(f"A{notice_row}:J{notice_row}")
        n = ws.cell(notice_row, 1,
                    value="⚠  Amber rows are DEMO data inserted so you can see the layout. "
                          "They will be replaced by real trades once the bot executes.")
        n.font      = Font(name=FONT_NAME, italic=True, color=C_AMBER_FG, size=9)
        n.fill      = fill(C_AMBER_BG)
        n.alignment = left()


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2 — Open Positions
# ─────────────────────────────────────────────────────────────────────────────

def build_open_positions(ws, positions: list):
    open_pos = [p for p in positions if p.get("status") == "open"]
    build_trade_log(ws, open_pos, sheet_title="Open Positions")
    # Override freeze + add extra column: Unrealised Expected Profit
    ws.freeze_panes = "A2"
    if not open_pos:
        ws.merge_cells("A5:P5")
        c = ws.cell(5, 1, value="No open positions at this time.")
        c.font      = Font(name=FONT_NAME, italic=True, color="888888", size=11)
        c.alignment = center()


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3 — P&L Summary
# ─────────────────────────────────────────────────────────────────────────────

def build_pnl_summary(ws, positions: list, sessions: list):
    ws.title = "P&L Summary"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 28, "B": 18, "C": 5, "D": 28, "E": 18})

    closed = [p for p in positions if p.get("status") == "closed"]
    open_p = [p for p in positions if p.get("status") == "open"]
    pnls   = [p.get("actual_profit") or 0.0 for p in closed]
    wins   = [p for p in closed if (p.get("actual_profit") or 0) > 0]
    costs  = [p.get("gross_cost") or 0.0 for p in closed]

    gross_pnl = round(sum(pnls), 4)
    est_fees  = sum(
        (calc_est_fee(p.get("kalshi_price"), p.get("kalshi_contracts")) or 0)
        for p in closed
    )
    net_pnl   = round(gross_pnl - est_fees, 4)
    win_rate  = (len(wins) / len(closed)) if closed else 0.0
    total_cost = sum(costs)
    deployed   = sum(p.get("gross_cost") or 0 for p in open_p)
    avg_edge   = (sum(p.get("actual_profit_pct") or 0 for p in closed) / len(closed)) if closed else 0
    avg_days   = (sum(p.get("days_to_resolution") or 0 for p in closed) / len(closed)) if closed else 0
    best  = max(pnls) if pnls else 0.0
    worst = min(pnls) if pnls else 0.0
    mdd   = max_drawdown(pnls)
    sh    = sharpe(pnls)

    # ── Banner ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    ws["A1"].value     = "  P&L Summary Dashboard"
    ws["A1"].font      = Font(name=FONT_NAME, bold=True, color=C_WHITE, size=13)
    ws["A1"].fill      = fill(C_DARK_BG)
    ws["A1"].alignment = left()
    ws.row_dimensions[1].height = 30

    def kpi_block(start_row, title, left_data: list, right_data: list):
        """Render a 2-column KPI block with a section title."""
        # Section header
        ws.merge_cells(f"A{start_row}:E{start_row}")
        h = ws.cell(start_row, 1, value=f"  {title}")
        h.font      = Font(name=FONT_NAME, bold=True, color=C_WHITE, size=10)
        h.fill      = fill(C_ACCENT)
        h.alignment = left()
        ws.row_dimensions[start_row].height = 20

        row = start_row + 1
        all_pairs = list(zip(left_data, right_data)) if len(left_data) >= len(right_data) \
            else list(zip(left_data, right_data[:len(left_data)]))
        # Pad if uneven
        max_len = max(len(left_data), len(right_data))
        lpad = left_data  + [("", "")] * (max_len - len(left_data))
        rpad = right_data + [("", "")] * (max_len - len(right_data))

        for (lk, lv), (rk, rv) in zip(lpad, rpad):
            ws.row_dimensions[row].height = 18
            alt = (row % 2 == 0)
            bg = C_LIGHT_GREY if alt else C_WHITE

            lk_c = ws.cell(row, 1, value=lk)
            lv_c = ws.cell(row, 2, value=lv)
            rk_c = ws.cell(row, 4, value=rk)
            rv_c = ws.cell(row, 5, value=rv)
            # Spacer col C
            ws.cell(row, 3).fill = fill(C_WHITE)

            for c in (lk_c, rk_c):
                c.font      = body_font(bold=True)
                c.fill      = fill(bg)
                c.alignment = left()
                c.border    = thin_border()
            for c in (lv_c, rv_c):
                c.font      = blue_font()
                c.fill      = fill(bg)
                c.alignment = right_align()
                c.border    = thin_border()

            # Colour P&L values
            if isinstance(lv, (int, float)) and lk and "P&L" in lk:
                lv_c.font = Font(name=FONT_NAME, color=(C_GREEN_FG if lv >= 0 else C_RED_FG),
                                 bold=True, size=10)
            if isinstance(rv, (int, float)) and rk and "P&L" in rk:
                rv_c.font = Font(name=FONT_NAME, color=(C_GREEN_FG if rv >= 0 else C_RED_FG),
                                 bold=True, size=10)

            row += 1
        return row + 1  # next free row with gap

    row = 2
    # ── Trade Performance ─────────────────────────────────────────────────────
    row = kpi_block(row, "Trade Performance",
        left_data=[
            ("Total Trades (closed)",  len(closed)),
            ("Wins",                   len(wins)),
            ("Losses",                 len(closed) - len(wins)),
            ("Win Rate",               f"{win_rate:.1%}"),
            ("Avg Edge %",             f"{avg_edge:.2%}"),
            ("Avg Days to Resolution", f"{avg_days:.1f}"),
        ],
        right_data=[
            ("Open Positions",         len(open_p)),
            ("Capital Deployed ($)",   f"${deployed:,.2f}"),
            ("Total Cost Deployed ($)",f"${total_cost:,.2f}"),
            ("Best Trade ($)",         f"${best:+,.2f}"),
            ("Worst Trade ($)",        f"${worst:+,.2f}"),
            ("Avg Profit / Trade ($)", f"${(gross_pnl/len(closed)):,.4f}" if closed else "$0.00"),
        ]
    )

    # ── P&L ───────────────────────────────────────────────────────────────────
    row = kpi_block(row, "P&L",
        left_data=[
            ("Gross P&L ($)",          gross_pnl),
            ("Est. Kalshi Fees ($)",   round(est_fees, 4)),
            ("Net P&L After Fees ($)", net_pnl),
        ],
        right_data=[
            ("Kalshi Fee Rate",        "7% of entry price × contracts"),
            ("Poly Fee Rate",          "2% taker (blocked US — N/A)"),
            ("Net / Gross Ratio",      f"{(net_pnl/gross_pnl):.1%}" if gross_pnl else "N/A"),
        ]
    )

    # ── Risk Metrics ──────────────────────────────────────────────────────────
    row = kpi_block(row, "Risk Metrics",
        left_data=[
            ("Sharpe Ratio",           sh),
            ("Max Drawdown ($)",       f"${mdd:,.4f}"),
            ("Calmar Ratio",           f"{(gross_pnl/mdd):.2f}" if mdd > 0 else "∞"),
        ],
        right_data=[
            ("Need ≥ 2 closed trades for full risk stats", ""),
            ("Sharpe > 1.5 = good; > 2 = excellent", ""),
            ("Max DD measures peak-to-trough loss", ""),
        ]
    )

    # ── Daily P&L table (embedded mini-table) ─────────────────────────────────
    ws.merge_cells(f"A{row}:E{row}")
    h = ws.cell(row, 1, value="  Daily P&L (last 30 days)")
    h.font      = Font(name=FONT_NAME, bold=True, color=C_WHITE, size=10)
    h.fill      = fill(C_ACCENT)
    h.alignment = left()
    ws.row_dimensions[row].height = 20
    row += 1

    # Sub-headers
    for col, hdr in zip([1, 2, 3, 4, 5],
                        ["Date", "Trades", "Daily P&L ($)", "Cumulative P&L ($)", ""]):
        c = ws.cell(row, col, value=hdr)
        c.font      = hdr_font()
        c.fill      = fill(C_MID_BG)
        c.alignment = center()
        c.border    = thin_border()
    ws.row_dimensions[row].height = 18
    row += 1

    # Group by day
    from collections import defaultdict
    by_day = defaultdict(list)
    for p in closed:
        if p.get("closed_at"):
            day = p["closed_at"][:10]
            by_day[day].append(p.get("actual_profit") or 0.0)

    cumulative = 0.0
    for day in sorted(by_day.keys()):
        daily = round(sum(by_day[day]), 4)
        cumulative = round(cumulative + daily, 4)
        alt = (row % 2 == 0)
        bg = C_LIGHT_GREY if alt else C_WHITE
        ws.row_dimensions[row].height = 16

        vals = [day, len(by_day[day]), daily, cumulative, ""]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row, col, value=val)
            c.fill      = fill(bg)
            c.font      = body_font()
            c.alignment = center() if col == 2 else right_align()
            c.border    = thin_border()
            if col in (3, 4):
                c.number_format = '$#,##0.0000;($#,##0.0000);"-"'
                c.font = Font(name=FONT_NAME,
                              color=(C_GREEN_FG if val >= 0 else C_RED_FG), size=10)
        row += 1

    if not by_day:
        ws.merge_cells(f"A{row}:D{row}")
        ws.cell(row, 1, value="No closed trades yet — daily P&L will appear here.").font = \
            Font(name=FONT_NAME, italic=True, color="888888", size=10)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 4 — Sessions
# ─────────────────────────────────────────────────────────────────────────────

SESSION_HEADERS = [
    "Session ID", "Mode", "Started At (UTC)", "Ended At (UTC)",
    "Duration (hrs)", "Trades", "Gross P&L ($)", "Est. Fees ($)",
    "Net P&L ($)", "Status",
]

def build_sessions(ws, sessions: list):
    ws.title = "Sessions"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    ws.merge_cells("A1:J1")
    ws["A1"].value     = "  Bot Session History"
    ws["A1"].font      = Font(name=FONT_NAME, bold=True, color=C_WHITE, size=12)
    ws["A1"].fill      = fill(C_DARK_BG)
    ws["A1"].alignment = left()
    ws.row_dimensions[1].height = 26

    for col_idx, hdr in enumerate(SESSION_HEADERS, 1):
        ws.cell(2, col_idx, value=hdr)
    style_header_row(ws, 2, bg=C_ACCENT)

    set_col_widths(ws, {
        "A": 12, "B": 8, "C": 18, "D": 18,
        "E": 14, "F": 9, "G": 14, "H": 12, "I": 12, "J": 10,
    })

    for row_idx, s in enumerate(sessions, start=3):
        alt = (row_idx % 2 == 0)
        style_data_row(ws, row_idx, alternate=alt)
        ws.row_dimensions[row_idx].height = 16

        hold_hrs = calc_hold_hours(s.get("started_at"), s.get("ended_at"))
        status   = "Active" if not s.get("ended_at") else "Ended"
        gross    = s.get("gross_pnl") or 0.0
        est_fee  = abs(gross) * 0.07 if gross else 0.0
        net      = (s.get("net_pnl") or gross) - est_fee

        row_data = [
            s.get("id"),
            (s.get("mode") or "paper").upper(),
            fmt_dt(s.get("started_at")),
            fmt_dt(s.get("ended_at")),
            hold_hrs,
            s.get("trades_executed") or 0,
            gross,
            round(est_fee, 4),
            round(net, 4),
            status,
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value=val)
            col_l = get_column_letter(col_idx)
            if col_l in ("G", "H", "I"):
                cell.number_format = '$#,##0.0000;($#,##0.0000);"-"'
                if col_l in ("G", "I") and isinstance(val, (int, float)):
                    cell.font = Font(name=FONT_NAME, size=10,
                                     color=(C_GREEN_FG if val >= 0 else C_RED_FG))

        # Status cell colour
        status_c = ws.cell(row_idx, 10)
        if status == "Active":
            status_c.fill = fill(C_GREEN_BG)
            status_c.font = Font(name=FONT_NAME, color=C_GREEN_FG, bold=True, size=10)
        else:
            status_c.fill = fill(C_LIGHT_GREY)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 5 — Scan Activity
# ─────────────────────────────────────────────────────────────────────────────

SCAN_HEADERS = [
    "Scan #", "Timestamp", "Kalshi Markets", "Poly/PredictIt Markets",
    "Candidates", "Opportunities", "Trades Fired", "Hit Rate",
]

def build_scan_activity(ws, scans: list):
    ws.title = "Scan Activity"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    ws.merge_cells("A1:H1")
    ws["A1"].value     = "  Scan Activity — parsed from bot log"
    ws["A1"].font      = Font(name=FONT_NAME, bold=True, color=C_WHITE, size=12)
    ws["A1"].fill      = fill(C_DARK_BG)
    ws["A1"].alignment = left()
    ws.row_dimensions[1].height = 26

    for col_idx, hdr in enumerate(SCAN_HEADERS, 1):
        ws.cell(2, col_idx, value=hdr)
    style_header_row(ws, 2, bg=C_ACCENT)

    set_col_widths(ws, {
        "A": 9, "B": 18, "C": 15, "D": 20,
        "E": 12, "F": 14, "G": 13, "H": 11,
    })

    total_scans = total_opps = total_trades = 0

    for row_idx, s in enumerate(scans[-500:], start=3):  # last 500 scans
        alt = (row_idx % 2 == 0)
        style_data_row(ws, row_idx, alternate=alt)
        ws.row_dimensions[row_idx].height = 15

        opps   = s.get("opportunities", 0)
        cands  = s.get("candidates", 0)
        trades = s.get("trades", 0)
        hit    = (trades / opps) if opps > 0 else 0.0

        row_data = [
            s.get("scan_num"), s.get("timestamp"),
            s.get("kalshi_mkts"), s.get("poly_mkts"),
            cands, opps, trades,
            f"{hit:.0%}" if opps > 0 else "—",
        ]

        for col_idx2, val in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx2, value=val)
            if col_idx2 == 6 and isinstance(val, int) and val > 0:
                cell.fill = fill(C_AMBER_BG)
                cell.font = Font(name=FONT_NAME, color=C_AMBER_FG, bold=True, size=10)
            if col_idx2 == 7 and isinstance(val, int) and val > 0:
                cell.fill = fill(C_GREEN_BG)
                cell.font = Font(name=FONT_NAME, color=C_GREEN_FG, bold=True, size=10)

        total_scans += 1
        total_opps  += opps
        total_trades += trades

    # Summary footer
    if scans:
        footer_row = len(scans[-500:]) + 4
        ws.row_dimensions[footer_row].height = 20
        ws.merge_cells(f"A{footer_row}:B{footer_row}")
        ws.cell(footer_row, 1, value="TOTALS").font = hdr_font()
        ws.cell(footer_row, 1).fill = fill(C_MID_BG)
        ws.cell(footer_row, 1).alignment = center()

        for col, val in [(5, total_scans), (6, total_opps), (7, total_trades)]:
            c = ws.cell(footer_row, col, value=val)
            c.font  = hdr_font()
            c.fill  = fill(C_MID_BG)
            c.alignment = center()

    if not scans:
        ws.merge_cells("A4:H4")
        ws.cell(4, 1, value="No scan log data found. Run the bot and re-generate.").font = \
            Font(name=FONT_NAME, italic=True, color="888888", size=10)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 6 — How to Refresh
# ─────────────────────────────────────────────────────────────────────────────

def build_instructions(ws):
    ws.title = "How to Refresh"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 4, "B": 60, "C": 4})

    ws.merge_cells("B1:B1")
    ws["B1"].value     = "How to Refresh This Report"
    ws["B1"].font      = Font(name=FONT_NAME, bold=True, color=C_WHITE, size=14)
    ws["B1"].fill      = fill(C_DARK_BG)
    ws["B1"].alignment = left()
    ws.row_dimensions[1].height = 32

    lines = [
        ("", ""),
        ("QUICK REFRESH", ""),
        ("", "From the arb-bot/ directory, run:"),
        ("", "    python scripts/generate_report.py"),
        ("", "Then open the new arb_trade_report.xlsx in your Projects folder."),
        ("", ""),
        ("CUSTOM OUTPUT PATH", ""),
        ("", "    python scripts/generate_report.py --output /path/to/my_report.xlsx"),
        ("", ""),
        ("DEMO DATA", ""),
        ("", "    python scripts/generate_report.py --demo"),
        ("", "Adds sample trades so you can preview every sheet before the bot fires."),
        ("", "Remove the flag once real trades appear in the DB."),
        ("", ""),
        ("DATA SOURCES", ""),
        ("", "  • Trade Log / Open Positions  →  data/arb_positions.db  (positions table)"),
        ("", "  • Sessions                    →  data/arb_positions.db  (sessions table)"),
        ("", "  • Scan Activity               →  data/paper_trade.log   (parsed line by line)"),
        ("", "                                    data/bot.log           (production mode)"),
        ("", ""),
        ("RECOMMENDED COLUMNS", ""),
        ("", "  Trade Log tracks the fields below — all recommended for arb monitoring:"),
        ("", ""),
        ("", "  ✓  Trade ID, Mode (paper/live), Status"),
        ("", "  ✓  Market Title (human-readable), Kalshi Ticker"),
        ("", "  ✓  Entry & Fill prices for both legs — shows actual slippage"),
        ("", "  ✓  Gross Cost — total capital at risk"),
        ("", "  ✓  Est. Kalshi Fee (7% × price × contracts)"),
        ("", "  ✓  Expected vs Actual Profit + P&L %"),
        ("", "  ✓  Net P&L After Fee — your real take-home"),
        ("", "  ✓  Hold Time (hours) — useful for spotting stuck positions"),
        ("", "  ✓  Days to Resolution — key for capital efficiency planning"),
        ("", "  ✓  Edge per Day, Annualised Return — normalise short vs long trades"),
        ("", "  ✓  Match Score + LLM Verified — quality signal for the cross-platform match"),
        ("", "  ✓  Close Reason — resolved_yes/no, manual, error"),
        ("", "  ✓  W/L — quick win/loss at a glance"),
        ("", ""),
        ("TAX NOTE", ""),
        ("", "  Kalshi contracts → US Section 1256 (60/40 long-/short-term treatment)"),
        ("", "  Polymarket      → Crypto/NFT rules — consult a tax professional"),
        ("", "  Run  python scripts/tax_export.py  (Phase 5) for Form 6781 prep"),
        ("", ""),
        ("TIPS", ""),
        ("", "  • Win Rate target: ≥ 85% paper before going live"),
        ("", "  • Sharpe target: ≥ 1.5 over 30+ trades"),
        ("", "  • Watch for Match Score < 0.80 trades — review manually"),
        ("", "  • Sort Trade Log by 'Actual P&L %' descending to find best markets"),
    ]

    for row_idx, (label, text) in enumerate(lines, start=2):
        ws.row_dimensions[row_idx].height = 17
        alt = (row_idx % 2 == 0)
        if label:
            ws.merge_cells(f"B{row_idx}:B{row_idx}")
            c = ws.cell(row_idx, 2, value=f"  {label}")
            c.font      = Font(name=FONT_NAME, bold=True, color=C_WHITE, size=10)
            c.fill      = fill(C_ACCENT)
            c.alignment = left()
        elif text:
            c = ws.cell(row_idx, 2, value=text)
            c.font      = body_font()
            c.fill      = fill(C_LIGHT_GREY if alt else C_WHITE)
            c.alignment = left()
            c.border    = thin_border()


# ─────────────────────────────────────────────────────────────────────────────
# Main builder
# ─────────────────────────────────────────────────────────────────────────────

def build_workbook(positions, sessions, scans, use_demo: bool) -> Workbook:
    if use_demo and not positions:
        print("  [info] No real trades found — injecting demo data.")
        positions = DEMO_POSITIONS
        if not sessions:
            sessions = DEMO_SESSIONS

    wb = Workbook()

    # Sheet 1 — Trade Log (default sheet)
    ws1 = wb.active
    ws1.title = "_placeholder"
    build_trade_log(ws1, positions, sheet_title="Trade Log")

    # Sheet 2 — Open Positions
    ws2 = wb.create_sheet("Open Positions")
    build_open_positions(ws2, positions)

    # Sheet 3 — P&L Summary
    ws3 = wb.create_sheet("P&L Summary")
    build_pnl_summary(ws3, positions, sessions)

    # Sheet 4 — Sessions
    ws4 = wb.create_sheet("Sessions")
    build_sessions(ws4, sessions)

    # Sheet 5 — Scan Activity
    ws5 = wb.create_sheet("Scan Activity")
    build_scan_activity(ws5, scans)

    # Sheet 6 — Instructions
    ws6 = wb.create_sheet("How to Refresh")
    build_instructions(ws6)

    # Make Trade Log the active sheet on open
    wb.active = ws1

    return wb


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Arb Bot — Excel Report Generator")
    parser.add_argument("--output", default=None,
                        help="Output path for the .xlsx file")
    parser.add_argument("--demo", action="store_true",
                        help="Inject demo trades when DB is empty")
    parser.add_argument("--db", default="data/arb_positions.db",
                        help="Path to SQLite DB (default: data/arb_positions.db)")
    parser.add_argument("--log", default=None,
                        help="Path to bot log for scan activity parsing")
    args = parser.parse_args()

    # Resolve paths relative to arb-bot/
    script_dir = os.path.dirname(os.path.abspath(__file__))
    bot_dir    = os.path.dirname(script_dir)
    db_path    = os.path.join(bot_dir, args.db)

    # Auto-detect log file
    if args.log:
        log_path = args.log
    else:
        for candidate in ["data/bot.log", "data/paper_trade.log",
                          "data/paper_trade_live.log"]:
            full = os.path.join(bot_dir, candidate)
            if os.path.exists(full) and os.path.getsize(full) > 0:
                log_path = full
                break
        else:
            log_path = None

    # Output path
    if args.output:
        out_path = args.output
    else:
        # Default: write to project root (one level above arb-bot/)
        project_dir = os.path.dirname(bot_dir)
        out_path    = os.path.join(project_dir, "arb_trade_report.xlsx")

    print(f"Arb Bot Report Generator")
    print(f"  DB   : {db_path}")
    print(f"  Log  : {log_path or 'not found'}")
    print(f"  Out  : {out_path}")
    print(f"  Demo : {args.demo}")
    print()

    # Load data
    print("Loading positions from DB...", end=" ")
    if os.path.exists(db_path):
        positions, fills, sessions = load_data(db_path)
        print(f"{len(positions)} positions, {len(sessions)} sessions")
    else:
        print("DB not found — using empty data")
        positions, fills, sessions = [], [], []

    print("Parsing scan log...", end=" ")
    scans = parse_scan_log(log_path) if log_path else []
    print(f"{len(scans)} scans")

    # Build workbook
    print("Building workbook...")
    wb = build_workbook(positions, sessions, scans, use_demo=args.demo)

    # Save
    os.makedirs(os.path.dirname(out_path) if os.path.dirname(out_path) else ".", exist_ok=True)
    wb.save(out_path)
    print(f"\n✓ Report saved: {out_path}")
    print(f"  Sheets: Trade Log | Open Positions | P&L Summary | Sessions | Scan Activity | How to Refresh")


if __name__ == "__main__":
    main()
